import os
import re
import json
import time
from datetime import datetime, time as dtime
from zoneinfo import ZoneInfo
from typing import Dict, List, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeoutError

# ==================
# Config (ENV)
# ==================
TZ = ZoneInfo(os.getenv("TZ", "Asia/Kuala_Lumpur"))

PRINTER_BASE_URL = os.getenv("PRINTER_BASE_URL", "http://192.168.1.50")
LOGIN_USER = os.getenv("LOGIN_USER", "admin")
LOGIN_PASS = os.getenv("LOGIN_PASS", "")  # blank password as requested

OUTPUT_ROOT = os.getenv("OUTPUT_ROOT", "/data/output")
HEADLESS = os.getenv("HEADLESS", "true").lower() == "true"

WORK_START = os.getenv("WORK_START", "08:00")  # 8am
WORK_END = os.getenv("WORK_END", "17:00")      # 5pm

POLL_WORK_SECONDS = int(os.getenv("POLL_WORK_SECONDS", "300"))   # 5 min
POLL_AFTER_SECONDS = int(os.getenv("POLL_AFTER_SECONDS", "1800")) # 30 min

# State file (for dedupe by Job ID per day)
STATE_DIR = os.getenv("STATE_DIR", "/data/state")

# =========================
# Helpers
# =========================
def now_local() -> datetime:
    return datetime.now(TZ)

def today_str() -> str:
    return now_local().strftime("%Y-%m-%d")

def ensure_dirs():
    os.makedirs(OUTPUT_ROOT, exist_ok=True)
    os.makedirs(STATE_DIR, exist_ok=True)

def parse_hhmm(s: str) -> dtime:
    hh, mm = s.split(":")
    return dtime(int(hh), int(mm))

WORK_START_T = parse_hhmm(WORK_START)
WORK_END_T = parse_hhmm(WORK_END)

def is_working_time(dt: datetime) -> bool:
    t = dt.time()
    return (t >= WORK_START_T) and (t <= WORK_END_T)

def get_poll_seconds(dt: datetime) -> int:
    return POLL_WORK_SECONDS if is_working_time(dt) else POLL_AFTER_SECONDS

def daily_paths() -> Tuple[str, str, str]:
    """
    Returns: (daily_folder, excel_path, state_path)
    """
    d = today_str()
    daily_folder = os.path.join(OUTPUT_ROOT, d)
    os.makedirs(daily_folder, exist_ok=True)

    excel_path = os.path.join(daily_folder, f"PrinterJobHistory_{d}.xlsx")
    state_path = os.path.join(STATE_DIR, f"state_{d}.json")
    return daily_folder, excel_path, state_path

def load_state(state_path: str) -> Dict:
    if os.path.exists(state_path):
        with open(state_path, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"seen_job_ids": []}

def save_state(state_path: str, state: Dict) -> None:
    with open(state_path, "w", encoding="utf-8") as f:
        json.dump(state, f, indent=2)

def safe_sheet_name(name: str) -> str:
    # Excel sheet constraints
    name = name.strip() if name else "UNKNOWN"
    name = re.sub(r"[\[\]\:\*\?\/\\]", "_", name)
    return name[:31] if len(name) > 31 else name

def ensure_workbook(excel_path: str) -> Workbook:
    if os.path.exists(excel_path):
        return load_workbook(excel_path)
    wb = Workbook()
    # Remove default sheet; we will create per-user sheets
    default = wb.active
    wb.remove(default)
    return wb

def ensure_user_sheet(wb: Workbook, user_id: str):
    sheet = safe_sheet_name(user_id)
    if sheet not in wb.sheetnames:
        ws = wb.create_sheet(sheet)
        ws.append(["UserID", "File Name", "Description", "Created at"])
        # Basic column sizing (optional but helps usability)
        for idx, width in enumerate([15, 45, 35, 22], start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width

def append_rows_by_user(excel_path: str, rows: List[Dict[str, str]]) -> int:
    wb = ensure_workbook(excel_path)
    added = 0

    for r in rows:
        user_id = r.get("UserID") or "UNKNOWN"
        ensure_user_sheet(wb, user_id)
        ws = wb[safe_sheet_name(user_id)]

        ws.append([
            user_id,
            r.get("FileName", ""),
            "",  # Description intentionally blank for staff input
            r.get("CreatedAt", ""),
        ])
        added += 1

    wb.save(excel_path)
    return added

# =========================
# Playwright navigation & scraping
# =========================
def login_and_go_history(page):
    """
    Flow:
    Login -> Print Job/Stored File -> Printer: Print Jobs -> Go to [Printer Job History] >>
    """
    # 1) Login page
    page.goto(PRINTER_BASE_URL, wait_until="networkidle")

    # These fields match your screenshot labels:
    # "Login User Name" and "Login Password"
    page.get_by_label("Login User Name").fill(LOGIN_USER)
    page.get_by_label("Login Password").fill(LOGIN_PASS)
    page.get_by_role("button", name="Login").click()

    # 2) Wait landing page (Web Image Monitor home)
    page.wait_for_load_state("networkidle")

    # 3) Left menu: Print Job/Stored File
    # Using text-based selection for robustness across firmware versions
    page.get_by_text("Print Job/Stored File", exact=False).click()
    page.wait_for_load_state("networkidle")

    # 4) Choose "Printer: Print Jobs"
    page.get_by_text("Printer: Print Jobs", exact=False).click()
    page.wait_for_load_state("networkidle")

    # 5) On Print Job List page: click "Go to [Printer Job History] >>"
    page.get_by_text("Go to [Printer Job History]", exact=False).click()
    page.wait_for_load_state("networkidle")

def scrape_printer_job_history(page) -> List[Dict[str, str]]:
    """
    Scrapes the table from Printer Job History.
    Expected columns in UI: ID | User Name | User ID | File Name | Status | Created At | Page(s) ...
    We extract: JobID, UserID, FileName, CreatedAt
    """
    # Table exists under the main content; take the first table after header.
    # We'll locate rows that contain numeric ID in first cell.
    table = page.locator("table").first
    table.wait_for(state="visible", timeout=15000)

    # Grab all rows
    rows = table.locator("tr")
    count = rows.count()

    results = []
    for i in range(count):
        tr = rows.nth(i)
        tds = tr.locator("td")
        if tds.count() < 6:
            continue

        # Extract text from key columns by position based on your screenshot:
        # 0=ID, 1=User Name, 2=User ID, 3=File Name, 4=Status, 5=Created At
        job_id = tds.nth(0).inner_text().strip()
        user_id = tds.nth(2).inner_text().strip()
        file_name = tds.nth(3).inner_text().strip()
        created_at = tds.nth(5).inner_text().strip()

        # Validate job_id looks like a number
        if not re.fullmatch(r"\d+", job_id):
            continue

        results.append({
            "JobID": job_id,
            "UserID": user_id if user_id and user_id != "---" else "UNKNOWN",
            "FileName": file_name,
            "CreatedAt": created_at,
        })

    return results

# =========================
# Main loop
# =========================
def run_once() -> Tuple[int, int, str]:
    """
    Returns (scraped_count, appended_count, excel_path)
    """
    _, excel_path, state_path = daily_paths()
    state = load_state(state_path)
    seen = set(state.get("seen_job_ids", []))

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=HEADLESS)
        context = browser.new_context(ignore_https_errors=True)  # Ricoh often uses self-signed cert
        page = context.new_page()

        try:
            login_and_go_history(page)
            scraped = scrape_printer_job_history(page)
        finally:
            context.close()
            browser.close()

    # Deduplicate by JobID (best control point)
    new_rows = []
    for r in scraped:
        if r["JobID"] not in seen:
            new_rows.append(r)
            seen.add(r["JobID"])

    appended = 0
    if new_rows:
        appended = append_rows_by_user(excel_path, new_rows)

    # Keep only last N IDs to control file size
    state["seen_job_ids"] = list(seen)[-10000:]
    save_state(state_path, state)

    return len(scraped), appended, excel_path


def main():
    ensure_dirs()
    print(f"[{now_local().isoformat()}] Ricoh Job History automation started ✅")

    while True:
        dt = now_local()
        poll = get_poll_seconds(dt)

        try:
            scraped, appended, excel_path = run_once()
            print(f"[{now_local().isoformat()}] Scraped={scraped} | New appended={appended} | File={excel_path}")
        except PWTimeoutError as e:
            print(f"[{now_local().isoformat()}] Playwright timeout: {e}")
        except Exception as e:
            print(f"[{now_local().isoformat()}] ERROR: {e}")

        time.sleep(poll)


if __name__ == "__main__":
    main()