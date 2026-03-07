import os
import re
from openpyxl import Workbook, load_workbook

def safe_sheet_name(name):
    name = (name or "UNKNOWN").strip() or "UNKNOWN"
    name = re.sub(r"[\[\]\:\*\?\/\\]", "_", name)
    return name[:31]

def ensure_workbook(path):
    if os.path.exists(path):
        return load_workbook(path)

    wb = Workbook()
    ws = wb.active
    ws.title = "TEMP"
    return wb

def append_jobs_by_user(path, jobs):
    """
    Append rows into per-user sheets.
    If jobs is empty, do nothing.
    """
    if not jobs:
        return 0

    wb = ensure_workbook(path)

    for job in jobs:
        sheet_name = safe_sheet_name(job.user_id)

        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            ws.append(["UserID", "File Name", "Description", "Created at"])

        ws = wb[sheet_name]
        ws.append([job.user_id, job.file_name, "", job.created_at])

    # Remove TEMP sheet only after real sheets exist
    if "TEMP" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["TEMP"]

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)
    return len(jobs)